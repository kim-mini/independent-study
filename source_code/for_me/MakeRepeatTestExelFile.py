# import pandas as pd
# df1 = pd.read_csv("C:/Users/82106/Desktop/Repeat Test/PIG_Vision_1.csv", "rb")
import openpyxl
import csv

ExelPath = "C:/Users/82106/Desktop/Repeat Test/fin_FIle.xlsx"
ExelPath1 = "C:/Users/82106/Desktop/Repeat Test/PIG_Vision_1.xlsx"
ExelPath2 = "C:/Users/82106/Desktop/Repeat Test/PIG_Vision_2.xlsx"
ExelPath3 = "C:/Users/82106/Desktop/Repeat Test/PIG_Vision_3.xlsx"
ExelPath4 = "C:/Users/82106/Desktop/Repeat Test/PIG_Vision_4.xlsx"
ExelPath5 = "C:/Users/82106/Desktop/Repeat Test/PIG_Vision_5.xlsx"

# Exel 파일 열기
load_wb = openpyxl.load_workbook(ExelPath)
finSheet = load_wb['Size']

load_wb1 = openpyxl.load_workbook(ExelPath1, data_only=True)
load_wb2 = openpyxl.load_workbook(ExelPath2, data_only=True)
load_wb3 = openpyxl.load_workbook(ExelPath3, data_only=True)
load_wb4 = openpyxl.load_workbook(ExelPath4, data_only=True)
load_wb5 = openpyxl.load_workbook(ExelPath5, data_only=True)

sheet1 = load_wb1['PIG_Vision_1']
sheet2 = load_wb2['PIG_Vision_2']
sheet3 = load_wb3['PIG_Vision_3']
sheet4 = load_wb4['PIG_Vision_4']
sheet5 = load_wb5['PIG_Vision_5']




rows = range(2, 2212) #47*47 기준
row_list = ['D', 'E']
for row in rows:
    finSheet['C'+str(row+1)] = sheet1['D'+str(row)].value
    finSheet['D'+str(row+1)] = sheet1['E'+str(row)].value

    finSheet['E'+str(row+1)] = sheet2['D'+str(row)].value
    finSheet['F'+str(row+1)] = sheet2['E'+str(row)].value

    finSheet['G' + str(row + 1)] = sheet3['D' + str(row)].value
    finSheet['H' + str(row + 1)] = sheet3['E' + str(row)].value

    finSheet['I' + str(row + 1)] = sheet4['D' + str(row)].value
    finSheet['J' + str(row + 1)] = sheet4['E' + str(row)].value

    finSheet['K' + str(row + 1)] = sheet5['D' + str(row)].value
    finSheet['L' + str(row + 1)] = sheet5['E' + str(row)].value

################SIZE

# Yag
# Ng_list = ['Black', 'White', 'Crack1_L', 'Crack1_R', 'Crack1_T', 'Crack1_B', 'Crack2_L', 'Crack2_R', 'Crack2_T', 'Crack2_B', 'Crack(Back)', 'Chip_L', 'Chip_R', 'Chip_T', 'Chip_B']
# ColIdx = ['I', 'M', 'Q', 'U', 'Y', 'AC', 'AG', 'AK', 'AO', 'AS', 'AW', 'BA', 'BE', 'BI', 'BM']
#Amber
Ng_list = ['Black', 'White', 'Crack1_L', 'Crack1_R', 'Crack1_T', 'Crack1_B', 'Crack2_L', 'Crack2_R', 'Crack2_T', 'Crack2_B', 'Chip_L', 'Chip_R', 'Chip_T', 'Chip_B']
ColIdx = ['I', 'M', 'Q', 'U', 'Y', 'AC', 'AG', 'AK', 'AO', 'AS', 'AW', 'BA', 'BE', 'BI']

idxCnt = 0
for NgName in Ng_list:
    finSheet = load_wb[NgName]
    for row in rows:
        finSheet['C'+str(row+1)] = sheet1[ColIdx[idxCnt]+str(row)].value
        finSheet['D'+str(row+1)] = sheet2[ColIdx[idxCnt]+str(row)].value
        finSheet['E'+str(row+1)] = sheet3[ColIdx[idxCnt]+str(row)].value
        finSheet['F'+str(row+1)] = sheet4[ColIdx[idxCnt]+str(row)].value
        finSheet['G'+str(row+1)] = sheet5[ColIdx[idxCnt]+str(row)].value
    idxCnt += 1


################




load_wb.save(ExelPath)
# #사용한 파일 닫기
# f1.close()
# f2.close()
